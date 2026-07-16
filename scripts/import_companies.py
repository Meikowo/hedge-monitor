#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_companies.py —— iFind 公司维表导入
==========================================
数据源：data/companies_ifind_YYYYMMDD.xlsx（iFind「全部A股」手动导出，季度刷新）。
列按表头文字匹配（不按位置），iFind 调整列序不影响本脚本。
自动清洗：页脚行（"数据来源：同花顺iFinD"）、空代码行。
企业性质规范化：中央企业→央企 / 地方国有企业→地方国企 / 私营→民企 /
              中外合资→外资 / 集体→集体 / 其余→其他（原值保留在 ent_nature_raw）。

用法：
  python scripts/import_companies.py --xlsx data/companies_ifind_20260711.xlsx --dry-run
  python scripts/import_companies.py --xlsx data/companies_ifind_20260711.xlsx
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import load_workbook

from common import ROOT, log, sb_upsert, snapshot_csv

# 表头文字（含子串即匹配）→ 内部字段
HEADER_MAP = [
    ("证券代码", "sec_code_full"),
    ("企业名称", "full_name"),
    ("企业中文简称", "short_name"),
    ("企业规模", "scale"),
    ("公司简介", "profile"),
    ("企业性质", "ent_nature_raw"),
    ("实际控制人", "actual_controller"),
    ("一级行业", "ind_l1"),
    ("二级行业", "ind_l2"),
    ("三级行业", "ind_l3"),
    ("省份", "province"),
    ("地级市", "city"),
]
ENT_TYPE_MAP = {"中央企业": "央企", "地方国有企业": "地方国企", "私营": "民企",
                "中外合资": "外资", "集体": "集体"}
CODE_RE = re.compile(r"^(\d{6})\.(SZ|SH|BJ)$")


def read_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c or "") for c in next(rows_iter)]
    col_idx: dict[str, int] = {}
    for key, field in HEADER_MAP:
        for i, h in enumerate(header):
            if key in h.replace("\n", ""):
                col_idx[field] = i
                break
    missing = [k for k, f in HEADER_MAP if f not in col_idx and f != "profile"]
    if missing:
        raise SystemExit(f"xlsx 缺少必需表头: {missing}（iFind 导出模板可能变了，需更新 HEADER_MAP）")

    out, dropped = [], 0
    for raw in rows_iter:
        rec = {f: (str(raw[i]).strip() if i < len(raw) and raw[i] is not None else None)
               for f, i in col_idx.items()}
        m = CODE_RE.match(rec.get("sec_code_full") or "")
        if not m:
            dropped += 1  # 页脚/空行
            continue
        nature = rec.get("ent_nature_raw") or ""
        out.append({
            "code": m.group(1),
            "market": m.group(2),
            "name": (rec.get("short_name") or "").split(";")[0] or None,
            "full_name": rec.get("full_name"),
            "ent_type": ENT_TYPE_MAP.get(nature, "其他"),
            "ent_nature_raw": nature or None,
            "actual_controller": rec.get("actual_controller"),
            "scale": rec.get("scale"),
            "ind_l1": rec.get("ind_l1"),
            "ind_l2": rec.get("ind_l2"),
            "ind_l3": rec.get("ind_l3"),
            "province": rec.get("province"),
            "city": rec.get("city"),
            "profile": rec.get("profile"),
            "source": path.name,
        })
    wb.close()
    log(f"读取 {len(out)} 家有效公司（丢弃非数据行 {dropped}）")
    return out


def main() -> None:
    ap = argparse.ArgumentParser(description="iFind 公司维表 → Supabase companies")
    ap.add_argument("--xlsx", default="data/companies_ifind_20260711.xlsx")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    path = (ROOT / args.xlsx) if not Path(args.xlsx).is_absolute() else Path(args.xlsx)
    if not path.exists():
        raise SystemExit(f"文件不存在: {path}")

    rows = read_xlsx(path)
    dup = len(rows) - len({r["code"] for r in rows})
    if dup:
        seen: set[str] = set()
        rows = [r for r in rows if not (r["code"] in seen or seen.add(r["code"]))]
        log(f"代码去重 {dup} 条（保留首条）")

    log("企业性质分布: " + ", ".join(f"{k}={v}" for k, v in
        Counter(r["ent_type"] for r in rows).most_common()))
    log("市场分布: " + ", ".join(f"{k}={v}" for k, v in
        Counter(r["market"] for r in rows).most_common()))
    log(f"一级行业数: {len({r['ind_l1'] for r in rows if r['ind_l1']})}, "
        f"三级行业数: {len({r['ind_l3'] for r in rows if r['ind_l3']})}")

    snapshot_csv("companies_import", [{k: v for k, v in r.items() if k != "profile"}
                                      for r in rows])
    if args.dry_run:
        log("dry-run：未写库")
        return
    n = sb_upsert("companies", rows, on_conflict="code")
    log(f"companies 已 upsert {n} 行")


if __name__ == "__main__":
    main()
